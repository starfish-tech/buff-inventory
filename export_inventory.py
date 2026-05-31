from __future__ import annotations

import argparse
import copy
import json
import logging
import math
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BUFF_API_URL = "https://buff.163.com/api/market/steam_inventory"
BUFF_INVENTORY_URL = "https://buff.163.com/market/steam_inventory"

HEADERS = [
    "饰品",
    "磨损",
    "买入价",
    "当前售价",
    "冷却时间",
    "利润",
    "利润率",
]


@dataclass
class InventoryRow:
    name: str = ""
    wear: str = ""
    buy_price: float | None = None
    current_price: float | None = None
    cooldown: str = ""


class BuffExportError(RuntimeError):
    pass


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise BuffExportError(
            f"Config file not found: {path}. Copy config.example.json to config.json first."
        )

    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise BuffExportError(f"Invalid JSON config: {exc}") from exc

    accounts = config.get("accounts")
    if not isinstance(accounts, list) or not accounts:
        raise BuffExportError("config.json must contain a non-empty accounts array.")

    for index, account in enumerate(accounts, start=1):
        for key in ("name", "steamid", "cookie"):
            if not account.get(key):
                raise BuffExportError(f"accounts[{index}] is missing required field: {key}")

    return config


def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"export_{datetime.now().strftime('%Y-%m-%d')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def clean_cookie(cookie: str) -> str:
    cookie = cookie.strip()
    if cookie.lower().startswith("cookie:"):
        cookie = cookie.split(":", 1)[1].strip()
    return cookie


def build_session(cookie: str, referer: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Connection": "keep-alive",
            "Cookie": clean_cookie(cookie),
            "Referer": referer,
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "X-Requested-With": "XMLHttpRequest",
        }
    )
    return session


def request_inventory_page(
    session: requests.Session,
    *,
    game: str,
    steamid: str,
    page_num: int,
    page_size: int,
    timeout: int,
) -> dict[str, Any]:
    params = {
        "game": game,
        "page_num": page_num,
        "page_size": page_size,
        "fold": "false",
        "search": "",
        "steamid": steamid,
        "state": "all",
    }
    response = session.get(BUFF_API_URL, params=params, timeout=timeout)

    try:
        payload = response.json()
    except ValueError as exc:
        raise BuffExportError(
            f"BUFF returned non-JSON response. HTTP {response.status_code}"
        ) from exc

    if response.status_code < 200 or response.status_code >= 300:
        raise BuffExportError(f"BUFF HTTP error: {response.status_code} {response.reason}")

    if payload.get("code") not in (None, "OK"):
        message = payload.get("msg") or payload.get("message") or payload.get("code")
        raise BuffExportError(f"BUFF API error: {message}")

    return payload


def get_payload_data(payload: dict[str, Any]) -> dict[str, Any]:
    data = payload.get("data", payload)
    if not isinstance(data, dict):
        raise BuffExportError("BUFF response data is not an object.")
    return data


def get_items(data: dict[str, Any]) -> list[dict[str, Any]]:
    items = data.get("items") or data.get("list") or data.get("assets")
    if items is None:
        return []
    if not isinstance(items, list):
        raise BuffExportError("BUFF response items is not an array.")
    return [item for item in items if isinstance(item, dict)]


def get_total_pages(data: dict[str, Any], page_size: int) -> int | None:
    for key in ("total_page", "total_pages", "page_count"):
        value = parse_number(data.get(key))
        if value:
            return int(value)

    total_count = parse_number(data.get("total_count") or data.get("total"))
    if total_count:
        return max(1, math.ceil(float(total_count) / page_size))

    return None


def merge_auxiliary_info(item: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    merged = copy.deepcopy(item)
    goods_id = str(first_non_empty(item, ["goods_id", "goodsId", "id"]) or "")
    asset_id = str(first_non_empty(item, ["assetid", "asset_id", "assetId"]) or "")

    for map_key, target_key, lookup_id in (
        ("goods_infos", "goods_info", goods_id),
        ("goods_info", "goods_info", goods_id),
        ("asset_infos", "asset_info", asset_id),
        ("asset_info", "asset_info", asset_id),
    ):
        source = data.get(map_key)
        if isinstance(source, dict) and lookup_id and isinstance(source.get(lookup_id), dict):
            merged.setdefault(target_key, {})
            if isinstance(merged[target_key], dict):
                merged[target_key] = {**source[lookup_id], **merged[target_key]}

    return merged


def first_non_empty(obj: dict[str, Any], paths: list[str]) -> Any:
    for path in paths:
        value = get_path(obj, path)
        if value not in (None, ""):
            return value
    return None


def get_path(obj: Any, path: str) -> Any:
    current = obj
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def recursive_key_values(obj: Any, key_pattern: re.Pattern[str]) -> list[Any]:
    values: list[Any] = []
    if isinstance(obj, dict):
        for key, value in obj.items():
            if key_pattern.search(str(key)):
                values.append(value)
            values.extend(recursive_key_values(value, key_pattern))
    elif isinstance(obj, list):
        for value in obj:
            values.extend(recursive_key_values(value, key_pattern))
    return values


def parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(",", "")
        normalized = normalized.replace("￥", "").replace("¥", "").strip()
        match = re.search(r"-?\d+(?:\.\d+)?", normalized)
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                return None
    return None


def normalize_excel_number(value: float | None) -> int | float | None:
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value


def parse_buy_price(item: dict[str, Any]) -> float | None:
    direct = first_non_empty(
        item,
        [
            "note",
            "remark",
            "comment",
            "memo",
            "asset_info.note",
            "asset_info.remark",
            "asset_info.comment",
            "asset_info.memo",
            "asset_info.info.note",
            "asset_info.info.remark",
            "asset_info.info.comment",
            "asset_info.info.memo",
        ],
    )
    parsed = parse_number(direct)
    if parsed is not None:
        return parsed

    note_values = recursive_key_values(
        item, re.compile(r"(note|remark|comment|memo|备注)", re.IGNORECASE)
    )
    for value in note_values:
        parsed = parse_number(value)
        if parsed is not None:
            return parsed

    return None


def parse_wear(item: dict[str, Any]) -> str:
    value = first_non_empty(
        item,
        [
            "paintwear",
            "paint_wear",
            "float",
            "wear",
            "asset_info.paintwear",
            "asset_info.paint_wear",
            "asset_info.float",
            "asset_info.wear",
            "asset_info.info.paintwear",
            "asset_info.info.paint_wear",
            "asset_info.info.float",
            "asset_info.info.wear",
        ],
    )
    if value in (None, ""):
        values = recursive_key_values(item, re.compile(r"(paint.*wear|float|wear)", re.IGNORECASE))
        value = next((entry for entry in values if parse_number(entry) is not None), None)

    if value in (None, ""):
        return ""
    return str(value).strip()


def parse_current_price(item: dict[str, Any]) -> float | None:
    value = first_non_empty(
        item,
        [
            "sell_min_price",
            "price",
            "market_price",
            "lowest_price",
            "goods_info.sell_min_price",
            "goods_info.price",
            "goods_info.market_price",
            "goods_info.lowest_price",
            "sell_order.price",
            "sell_order.sell_min_price",
        ],
    )
    parsed = parse_number(value)
    if parsed is not None:
        return parsed

    price_values = recursive_key_values(
        item, re.compile(r"(sell_min_price|lowest_price|market_price)", re.IGNORECASE)
    )
    for candidate in price_values:
        parsed = parse_number(candidate)
        if parsed is not None:
            return parsed
    return None


def contains_cjk(value: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in value)


def parse_name(item: dict[str, Any]) -> str:
    paths = [
        "goods_info.name",
        "asset_info.name",
        "asset_info.info.name",
        "name",
        "goods_info.localized_name",
        "asset_info.localized_name",
        "asset_info.info.localized_name",
        "localized_name",
        "goods_info.short_name",
        "asset_info.info.short_name",
        "goods_info.market_name",
        "asset_info.market_name",
        "asset_info.info.market_name",
        "market_name",
        "goods_info.market_hash_name",
        "asset_info.market_hash_name",
        "asset_info.info.market_hash_name",
        "market_hash_name",
    ]
    candidates = [str(value).strip() for value in (get_path(item, path) for path in paths) if value not in (None, "")]
    for candidate in candidates:
        if contains_cjk(candidate):
            return candidate
    return candidates[0] if candidates else ""


def parse_cooldown(item: dict[str, Any]) -> str:
    value = first_non_empty(
        item,
        [
            "cooldown",
            "trade_cooldown",
            "tradable_cooldown",
            "tradable_cooldown_text",
            "trade_cd",
            "tradable_unfrozen_time",
            "unfrozen_time",
            "asset_info.cooldown",
            "asset_info.trade_cooldown",
            "asset_info.tradable_cooldown",
            "asset_info.tradable_cooldown_text",
            "asset_info.trade_cd",
            "asset_info.tradable_unfrozen_time",
            "asset_info.unfrozen_time",
            "asset_info.info.tradable_unfrozen_time",
            "asset_info.info.unfrozen_time",
        ],
    )

    parsed = cooldown_to_text(value)
    if parsed:
        return parsed

    values = recursive_key_values(
        item, re.compile(r"(cooldown|trade_cd|unfrozen|tradable.*time)", re.IGNORECASE)
    )
    for candidate in values:
        parsed = cooldown_to_text(candidate)
        if parsed:
            return parsed

    return ""


def cooldown_to_text(value: Any) -> str:
    if value in (None, "", 0, "0"):
        return ""

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return ""
        if re.search(r"\d+\s*天", text):
            number = re.search(r"\d+", text)
            return f"{number.group(0)}天" if number else text
        if "可交易" in text or "无" == text:
            return ""

    number = parse_number(value)
    if number is None:
        return str(value).strip() if isinstance(value, str) else ""

    now = time.time()
    timestamp = number / 1000 if number > 10_000_000_000 else number
    if timestamp > now + 3600:
        days = math.ceil((timestamp - now) / 86400)
        return f"{days}天"

    if 0 < number <= 30:
        return f"{int(math.ceil(number))}天"

    return ""


def normalize_item(item: dict[str, Any], data: dict[str, Any]) -> InventoryRow:
    merged = merge_auxiliary_info(item, data)
    return InventoryRow(
        name=parse_name(merged),
        wear=parse_wear(merged),
        buy_price=parse_buy_price(merged),
        current_price=parse_current_price(merged),
        cooldown=parse_cooldown(merged),
    )


def save_debug_payload(log_dir: Path, account_name: str, payload: dict[str, Any]) -> None:
    safe_name = sanitize_filename(account_name)
    debug_path = log_dir / f"{safe_name}_first_page.json"
    debug_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_account_rows(
    account: dict[str, Any],
    *,
    game: str,
    page_size: int,
    timeout: int,
    log_dir: Path,
    save_debug: bool,
) -> list[InventoryRow]:
    name = str(account["name"])
    steamid = str(account["steamid"])
    referer = (
        f"{BUFF_INVENTORY_URL}?game={game}#page_num=1&page_size={page_size}"
        f"&fold=false&search=&steamid={steamid}&state=all"
    )
    session = build_session(account["cookie"], referer)

    logging.info("Fetching account=%s steamid=%s", name, steamid)
    rows: list[InventoryRow] = []
    page_num = 1
    total_pages: int | None = None

    while True:
        payload = request_inventory_page(
            session,
            game=game,
            steamid=steamid,
            page_num=page_num,
            page_size=page_size,
            timeout=timeout,
        )
        data = get_payload_data(payload)
        items = get_items(data)

        if page_num == 1:
            total_pages = get_total_pages(data, page_size)
            if save_debug:
                save_debug_payload(log_dir, name, payload)

        rows.extend(normalize_item(item, data) for item in items)
        logging.info(
            "Fetched account=%s page=%s items=%s total_rows=%s",
            name,
            page_num,
            len(items),
            len(rows),
        )

        if total_pages is not None and page_num >= total_pages:
            break
        if total_pages is None and len(items) < page_size:
            break
        if not items:
            break

        page_num += 1
        time.sleep(0.5)

    return rows


def sanitize_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", value).strip()
    return (cleaned or "account")[:31]


def sanitize_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("_") or "account"


def unique_output_path(output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    date_part = datetime.now().strftime("%Y-%m-%d")
    path = output_dir / f"buff_inventory_{date_part}.xlsx"
    if not path.exists():
        return path
    return output_dir / f"buff_inventory_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"


def write_workbook(account_rows: dict[str, list[InventoryRow]], output_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_sheet_names: set[str] = set()
    for account_name, rows in account_rows.items():
        sheet_name = sanitize_sheet_name(account_name)
        base_name = sheet_name
        suffix = 2
        while sheet_name in used_sheet_names:
            suffix_text = f"_{suffix}"
            sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_sheet_names.add(sheet_name)

        ws = workbook.create_sheet(sheet_name)
        ws.append(HEADERS)

        for index, row in enumerate(reversed(rows), start=2):
            ws.cell(row=index, column=1, value=row.name)
            ws.cell(row=index, column=2, value=row.wear)
            ws.cell(row=index, column=3, value=normalize_excel_number(row.buy_price))
            ws.cell(row=index, column=4, value=normalize_excel_number(row.current_price))
            ws.cell(row=index, column=5, value=row.cooldown)
            ws.cell(row=index, column=6, value=f'=IF(OR(C{index}="",D{index}=""),"",D{index}-C{index})')
            ws.cell(row=index, column=7, value=f'=IF(OR(C{index}="",F{index}=""),"",F{index}/C{index})')

        format_sheet(ws)

    workbook.save(output_path)


def format_sheet(ws: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = [42, 22, 14, 14, 12, 14, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in range(2, ws.max_row + 1):
        for column in (3, 4, 6):
            ws.cell(row=row, column=column).number_format = "General"
        ws.cell(row=row, column=7).number_format = "0.00%"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export BUFF CS inventory to Excel.")
    parser.add_argument("--config", default="config.json", help="Path to config JSON.")
    parser.add_argument("--debug-json", action="store_true", help="Save first page JSON per account.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(__file__).resolve().parent
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = root / config_path

    try:
        config = load_config(config_path)
        log_dir = root / "logs"
        setup_logging(log_dir)

        game = str(config.get("game", "csgo"))
        page_size = int(config.get("page_size", 50))
        timeout = int(config.get("timeout", 30))
        output_dir = root / str(config.get("output_dir", "exports"))

        account_rows: dict[str, list[InventoryRow]] = {}
        for account in config["accounts"]:
            rows = fetch_account_rows(
                account,
                game=game,
                page_size=page_size,
                timeout=timeout,
                log_dir=log_dir,
                save_debug=args.debug_json,
            )
            account_rows[str(account["name"])] = rows

        output_path = unique_output_path(output_dir)
        write_workbook(account_rows, output_path)
        logging.info("Exported workbook: %s", output_path)
        print(output_path)
        return 0
    except Exception as exc:
        logging.exception("Export failed")
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
